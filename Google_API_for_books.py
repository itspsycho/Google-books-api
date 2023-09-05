import requests
import openpyxl
import tkinter as tk

def get_book_info(book_name):
    url = "https://www.googleapis.com/books/v1/volumes"
    params = {
        "q": f"intitle:{book_name}",
        "maxResults": 1
    }

    response = requests.get(url, params=params)
    data = response.json()

    if "items" in data:
        item = data["items"][0]
        volume_info = item.get("volumeInfo", {})
        title = volume_info.get("title", "")
        authors = volume_info.get("authors", [])
        identifiers = volume_info.get("industryIdentifiers", [])
        isbn = ""
        for identifier in identifiers:
            if identifier.get("type", "") == "ISBN_13":
                isbn = identifier.get("identifier", "")
                break

        book_id = item.get("id", "")
        cover_link = f"https://books.google.com/books/publisher/content/images/frontcover/{book_id}?fife=w1920-h1080&source=gbs_api"
        
        description = volume_info.get("description", "")
        
        return title, authors, cover_link, isbn, description

    return None, None, None, None, None

# Specify the path to the Excel file (Change this to your excel file path)
excel_file = r"C:\Users\user\Desktop\Book.xlsx"

# Load Excel file
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get the total number of books
total_books = sheet.max_row - 1

# Initialize counter
counter = 0

# Create a Tkinter window
window = tk.Tk()
window.title("Progress")

# Create a Label widget to display progress message
progress_label = tk.Label(window, text="")
progress_label.pack(pady=10)

# Iterate through rows
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
    book_name = row[0]
    title, authors, cover_link, isbn, description = get_book_info(book_name)

    # Update values in the row
    sheet.cell(row=row_index, column=2).value = isbn
    sheet.cell(row=row_index, column=3).value = title
    sheet.cell(row=row_index, column=4).value = ", ".join(authors) if authors else ""
    sheet.cell(row=row_index, column=5).value = cover_link
    sheet.cell(row=row_index, column=6).value = description
    
    # Increment counter
    counter += 1

    # Update progress message
    progress_message = f"Processed {counter} out of {total_books} books."
    progress_label.config(text=progress_message)
    window.update()

# Save the modified Excel file
workbook.save(excel_file)

# Close the Tkinter window
window.destroy()
